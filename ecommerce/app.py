# importing openpyxl
import openpyxl as xl
from  openpyxl.chart import BarChart ,Reference
wb =xl.load_workbook("transactions.xlsx") # loading excel file
sheet  = wb['Sheet1']  #getting sheet 1

cell =sheet['a1'] # getting first row and column
celle = sheet.cell(1,1)  # same as getting first row and column
print(celle.value)
print(cell.value ," same thing")
print("number of rows:"  ,sheet.max_row)  # to find out how many rows


for number in range(2 ,sheet.max_row + 1): # ignore heading -column and start from the second row
   # print(number)
    cell  = sheet.cell(number ,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(number ,4)
    corrected_price_cell.value =corrected_price

    print(".................................")
    print("correct price:", corrected_price)
    print(cell , ": " ,cell.value)
    print(cell.value)

# selecting a range of values using Reference
values = Reference(sheet , min_row=2 , max_row= sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart ,'e2')

wb.save('transactions2.xlsx')

